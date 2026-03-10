from flask import Blueprint, request, jsonify, send_file
import os
from openpyxl import load_workbook
from models import db, Provider, Rate, RatesFile

rates_bp = Blueprint("rates_bp", __name__)

IN_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), "in")


def _normalize_scope(scope_raw: str) -> str:
    s = (scope_raw or "").strip()

    if not s:
        raise ValueError("Scope is required")

    if s.upper() == "ALL":
        return "ALL"

    if not s.isdigit():
        raise ValueError(f"Scope must be 'ALL' or provider id. Got '{s}'")

    provider_id = int(s)

    if Provider.query.get(provider_id) is None:
        raise ValueError(f"Provider id in Scope does not exist: {provider_id}")

    return str(provider_id)


def _read_rates_excel(filename: str):
    safe_name = os.path.basename(filename)
    path = os.path.join(IN_DIR, safe_name)

    if not os.path.exists(path):
        raise FileNotFoundError(f"file not found in /in: {safe_name}")

    wb = load_workbook(path, data_only=True)
    ws = wb.active

    # read header row
    header = []
    for cell in ws[1]:
        header.append(str(cell.value).strip() if cell.value is not None else "")

    col_idx = {name: i for i, name in enumerate(header)}
    required = ["Product", "Rate", "Scope"]

    for col in required:
        if col not in col_idx:
            raise ValueError(f"missing column '{col}' in excel")

    rows = []

    for r in ws.iter_rows(min_row=2, values_only=True):
        product = r[col_idx["Product"]]
        rate_val = r[col_idx["Rate"]]
        scope = r[col_idx["Scope"]]

        product = str(product).strip() if product is not None else ""

        if not product:
            continue

        try:
            rate_int = int(rate_val)
        except Exception:
            raise ValueError(f"invalid Rate for product '{product}'")

        if rate_int < 0:
            raise ValueError(f"Rate must be non-negative for product '{product}'")

        scope_norm = _normalize_scope(str(scope) if scope is not None else "")

        rows.append({
            "product_id": product,
            "rate": rate_int,
            "scope": scope_norm
        })

    if not rows:
        raise ValueError("No valid rate rows found in excel")

    return rows, safe_name


# ---- added non-required endpoint ---- #
# Lists available files in /in for the frontend dropdown
@rates_bp.route("/files", methods=["GET"])
def list_files():
    files = [f for f in os.listdir(IN_DIR) if f.endswith((".xlsx", ".xls"))]
    return jsonify(files), 200
# ---- end non-required endpoint ---- #

@rates_bp.route("/rates", methods=["POST"])
def post_rates():
    """
    file = excel filename that already exists in /in
    Reads Product, Rate, Scope from excel and updates Rates table.
    Also saves the latest uploaded file name for GET /rates.
    """
    data = request.get_json(silent=True) or {}
    filename = data.get("file") or request.form.get("file")

    if not filename:
        return jsonify({"error": "file is required (name of excel file in /in)"}), 400

    try:
        rows, safe_name = _read_rates_excel(filename)

        inserted = 0
        updated = 0

        for row in rows:
            existing = Rate.query.filter_by(
                product_id=row["product_id"],
                scope=row["scope"]
            ).first()

            if existing:
                if existing.rate != row["rate"]:
                    existing.rate = row["rate"]
                    updated += 1
            else:
                new_rate = Rate(
                    product_id=row["product_id"],
                    scope=row["scope"],
                    rate=row["rate"]
                )
                db.session.add(new_rate)
                inserted += 1

        # save the latest uploaded file name
        latest_file = RatesFile.query.first()

        if latest_file:
            latest_file.filename = safe_name
        else:
            db.session.add(RatesFile(filename=safe_name))

        db.session.commit()

        return jsonify({
            "status": "OK",
            "file": safe_name,
            "rows": len(rows),
            "inserted": inserted,
            "updated": updated
        }), 200

    except FileNotFoundError as e:
        return jsonify({"error": "file not found", "details": str(e)}), 404

    except ValueError as e:
        return jsonify({"error": "bad input", "details": str(e)}), 400

    except Exception as e:
        db.session.rollback()
        return jsonify({"error": "server error", "details": str(e)}), 500


@rates_bp.route("/rates", methods=["GET"])
def get_rates():
    """
    Returns the same excel file that was last uploaded using POST /rates
    """
    try:
        latest_file = RatesFile.query.first()

        if latest_file is None:
            return jsonify({"error": "no rates file uploaded yet"}), 404

        safe_name = os.path.basename(latest_file.filename)
        path = os.path.join(IN_DIR, safe_name)

        if not os.path.exists(path):
            return jsonify({
                "error": "rates file record exists but file is missing",
                "file": safe_name
            }), 404

        return send_file(
            path,
            as_attachment=True,
            download_name=safe_name,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        return jsonify({"error": "server error", "details": str(e)}), 500